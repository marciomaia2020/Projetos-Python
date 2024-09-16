"""
#Este código copia as informações criando uma planilha nova chamada arquivos_lista.xlsx

import os
import pandas as pd

# Defina o caminho para a pasta que você deseja acessar
pasta = r"D:\MEUSSITESEPROJETOS\Sites\portifolioDoMaia\docs\python\pdf\separados"

# Defina o caminho completo onde você deseja salvar o arquivo Excel
caminho_arquivo_excel = r"D:\MEUSSITESEPROJETOS\Sites\portifolioDoMaia\docs\python\pdf\arquivos_lista.xlsx"

# Verifique se o caminho da pasta existe
if not os.path.exists(pasta):
    print(f"O caminho especificado não foi encontrado: {pasta}")
else:
    # Lista para armazenar os nomes e extensões dos arquivos
    lista_arquivos = []

    # Itera sobre os arquivos na pasta
    for arquivo in os.listdir(pasta):
        # Verifica se é um arquivo (ignora pastas)
        if os.path.isfile(os.path.join(pasta, arquivo)):
            nome, extensao = os.path.splitext(arquivo)
            lista_arquivos.append({"Nome": nome, "Extensão": extensao})

    # Cria um DataFrame com a lista de arquivos
    df = pd.DataFrame(lista_arquivos)

    # Salva o DataFrame em um arquivo Excel no caminho especificado
    df.to_excel(caminho_arquivo_excel, index=False)

    print(f"Arquivo Excel gerado com sucesso em: {caminho_arquivo_excel}")
"""

#Este código copia as informações diretamente para uma guia dentro de uma planilha ja especificada
import os
import pandas as pd
from openpyxl import load_workbook, Workbook

# Defina o caminho para a pasta que você deseja acessar
pasta = r"D:\MEUSSITESEPROJETOS\Sites\portifolioDoMaia\docs\python\pdf\separados"

# Defina o caminho completo onde está o arquivo Excel
caminho_arquivo_excel = r"D:\MEUSSITESEPROJETOS\Sites\portifolioDoMaia\docs\python\html\NumberPages.xlsx"

# Verifique se o caminho da pasta existe
if not os.path.exists(pasta):
    print(f"O caminho especificado não foi encontrado: {pasta}")
else:
    # Lista para armazenar os nomes e extensões dos arquivos
    lista_arquivos = []

    # Itera sobre os arquivos na pasta
    for arquivo in os.listdir(pasta):
        # Verifica se é um arquivo (ignora pastas)
        if os.path.isfile(os.path.join(pasta, arquivo)):
            nome, extensao = os.path.splitext(arquivo)
            lista_arquivos.append({"Nome": nome, "Extensão": extensao})

    # Cria um DataFrame com a lista de arquivos
    df = pd.DataFrame(lista_arquivos)

    try:
        # Tenta carregar o workbook existente
        workbook = load_workbook(caminho_arquivo_excel)

        # Verifica se a planilha 'Plan2' já existe, se não, cria uma nova
        if 'Plan2' in workbook.sheetnames:
            sheet = workbook['Plan2']
        else:
            sheet = workbook.create_sheet('Plan2')

        # Usa o ExcelWriter para modificar a planilha existente
        with pd.ExcelWriter(caminho_arquivo_excel, engine='openpyxl', mode='a') as writer:
            writer.book = workbook
            writer.sheets = {ws.title: ws for ws in workbook.worksheets}
            
            # Escreve o DataFrame na planilha 'Plan2', começando na célula A1
            df.to_excel(writer, sheet_name='Plan2', index=False)

        print(f"Dados inseridos com sucesso na Plan2 do arquivo: {caminho_arquivo_excel}")

    except Exception as e:
        print(f"Erro ao processar o arquivo Excel: {e}")
        print("Criando um novo arquivo Excel...")

        # Cria um novo workbook e uma planilha
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Plan1"

        # Adiciona os dados à nova planilha 'Plan1'
        for r_idx, (index, row) in enumerate(df.iterrows(), start=1):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        # Salva o novo arquivo Excel
        novo_caminho_excel = r"D:\MEUSSITESEPROJETOS\Sites\portifolioDoMaia\docs\python\html\NovoNumberPages.xlsx"
        workbook.save(novo_caminho_excel)
        print(f"Novo arquivo Excel criado com sucesso em: {novo_caminho_excel}")
