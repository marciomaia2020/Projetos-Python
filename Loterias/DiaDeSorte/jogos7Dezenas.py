
#COLUNA AJUSTÁVEL AUTOMATICAMENTE
import pandas as pd
import numpy as np
from itertools import combinations
from openpyxl import load_workbook

# Função para solicitar o mês ao usuário
def escolher_mes():
    print("Escolha um mês (1 a 12):")
    mes = int(input())
    
    # Verificar se o mês é válido
    if 1 <= mes <= 12:
        return mes
    else:
        print("Mês inválido! Por favor, escolha um número de 1 a 12.")
        return escolher_mes()  # Chama a função recursivamente se o mês for inválido

# Solicitar ao usuário que escolha um mês
mes_selecionado = escolher_mes()

# Gerar todas as combinações de 7 dezenas entre 31
dezenas = list(range(1, 32))  # Dezenas de 1 a 31
jogos = list(combinations(dezenas, 7))

# Número de jogos por planilha
num_jogos_por_planilha = 524286  # Ajuste conforme necessário
# FICA FALTANDO UM JOGO NO FINAL

# Selecionar jogos para a primeira planilha
selected_jogos1 = np.random.choice(len(jogos), num_jogos_por_planilha, replace=False)
df_jogos1 = pd.DataFrame([list(jogos[i]) + [''] + [mes_selecionado] for i in selected_jogos1])  # Adiciona uma coluna vazia antes do mês

# Definir o caminho absoluto para salvar as planilhas
caminho_arquivo1 = r'D:\MEUSSITESEPROJETOS\Projetos\Projetos-Python\Loterias\DiaDeSorte\jogos_dia_de_sorte_1.xlsx'
caminho_arquivo2 = r'D:\MEUSSITESEPROJETOS\Projetos\Projetos-Python\Loterias\DiaDeSorte\jogos_dia_de_sorte_2.xlsx'

# Salvar a primeira planilha
df_jogos1.to_excel(caminho_arquivo1, index=False, header=False)

# Ajustar a largura das colunas
wb1 = load_workbook(caminho_arquivo1)
ws1 = wb1.active
for column in ws1.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)  # Adiciona um pouco de espaço
    ws1.column_dimensions[column[0].column_letter].width = adjusted_width
wb1.save(caminho_arquivo1)

# Selecionar jogos para a segunda planilha
selected_jogos2 = np.random.choice(len(jogos), num_jogos_por_planilha, replace=False)
df_jogos2 = pd.DataFrame([list(jogos[i]) + [''] + [mes_selecionado] for i in selected_jogos2])  # Adiciona uma coluna vazia antes do mês

# Salvar a segunda planilha
df_jogos2.to_excel(caminho_arquivo2, index=False, header=False)

# Ajustar a largura das colunas na segunda planilha
wb2 = load_workbook(caminho_arquivo2)
ws2 = wb2.active
for column in ws2.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)  # Adiciona um pouco de espaço
    ws2.column_dimensions[column[0].column_letter].width = adjusted_width
wb2.save(caminho_arquivo2)

# Avisar que a criação das planilhas foi concluída
print("As planilhas foram geradas com sucesso!")
