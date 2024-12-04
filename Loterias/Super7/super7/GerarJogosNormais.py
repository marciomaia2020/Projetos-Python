import os
import pandas as pd
import numpy as np
from itertools import product
from openpyxl import load_workbook

# Função para verificar e criar o diretório se não existir
def verificar_criar_diretorio(caminho_diretorio):
    if not os.path.exists(caminho_diretorio):
        print(f"Diretório '{caminho_diretorio}' não encontrado. Criando o diretório...")
        os.makedirs(caminho_diretorio)
    else:
        print(f"Diretório '{caminho_diretorio}' já existe.")

# Gerar todas as combinações possíveis para o Super Sete (7 colunas, números de 0 a 9)
print("Gerando todas as combinações possíveis para o Super Sete...")
colunas = list(range(10))  # Números de 0 a 9
jogos = list(product(colunas, repeat=7))  # Combinações de 7 números (1 por coluna)
print(f"Total de combinações geradas: {len(jogos)}")

# Número de jogos por planilha
num_jogos_por_planilha = 524286  # Ajuste conforme necessário
print(f"Selecionando {num_jogos_por_planilha} jogos para cada planilha...")

# Caminho absoluto para salvar as planilhas
caminho_arquivo1 = r'D:\MEUSSITESEPROJETOS\Projetos\Projetos-Python\Loterias\SuperSete\jogos_super_sete_1.xlsx'
caminho_arquivo2 = r'D:\MEUSSITESEPROJETOS\Projetos\Projetos-Python\Loterias\SuperSete\jogos_super_sete_2.xlsx'

# Verificar se os diretórios existem, se não, criar
verificar_criar_diretorio(os.path.dirname(caminho_arquivo1))
verificar_criar_diretorio(os.path.dirname(caminho_arquivo2))

# Selecionar jogos para a primeira planilha
print("Selecionando jogos para a primeira planilha...")
selected_jogos1 = np.random.choice(len(jogos), num_jogos_por_planilha, replace=False)
df_jogos1 = pd.DataFrame([list(jogos[i]) for i in selected_jogos1])

# Salvar a primeira planilha
print(f"Salvando a primeira planilha em: {caminho_arquivo1}")
df_jogos1.to_excel(caminho_arquivo1, index=False, header=False)

# Ajustar a largura das colunas na primeira planilha
print("Ajustando a largura das colunas na primeira planilha...")
wb1 = load_workbook(caminho_arquivo1)
ws1 = wb1.active
for column in ws1.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)  # Adiciona um pouco de espaço
    ws1.column_dimensions[column[0].column_letter].width = adjusted_width
wb1.save(caminho_arquivo1)

# Selecionar jogos para a segunda planilha
print("Selecionando jogos para a segunda planilha...")
selected_jogos2 = np.random.choice(len(jogos), num_jogos_por_planilha, replace=False)
df_jogos2 = pd.DataFrame([list(jogos[i]) for i in selected_jogos2])

# Salvar a segunda planilha
print(f"Salvando a segunda planilha em: {caminho_arquivo2}")
df_jogos2.to_excel(caminho_arquivo2, index=False, header=False)

# Ajustar a largura das colunas na segunda planilha
print("Ajustando a largura das colunas na segunda planilha...")
wb2 = load_workbook(caminho_arquivo2)
ws2 = wb2.active
for column in ws2.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)  # Adiciona um pouco de espaço
    ws2.column_dimensions[column[0].column_letter].width = adjusted_width
wb2.save(caminho_arquivo2)

# Avisar que a criação das planilhas foi concluída
print("As planilhas do Super Sete foram geradas e ajustadas com sucesso!")
