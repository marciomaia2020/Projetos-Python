"""
1 - Entrar na plnilha e EXTRAIR O CPF do cliente
2 - Acessar o site https://consultcpf-devaprender.netlify.app e INSERIR O CPF para PESQUISAR O STATUS DO PAGAMENTO daquele cliente (cpf)
3 - Verificar se está em dia ou atrasado
4 - Se estiver "EM DIA", pegar a DATA e METODO de pagamento 
5 - Caso contrário (se estiver atrasado), colocar o status como pendente
6 - Inserir essas novas informações (nome, valor, cpf e vencimento - status) e caso esteja em dia, 
DATA DE PAGAMENTO, E METODO PAGAMENTO (cartão ou boleto) e, uma nova planilha
"""

# Instalar a dependencia 
# pip install openpyxl selenium (windows)
# pip3 install openpyxl selenium (mac ou linux)

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


# 1 - Entrar na planilha e EXTRAIR O CPF do cliente
planilha_clientes = openpyxl.load_workbook('D:\\MEUSSITESEPROJETOS\\Projetos\\Projetos-Python\\Projects4\\dados_clientes.xlsx')
pagina_clientes = planilha_clientes['Sheet1']

# 2 - Acessar o site e INSERIR O CPF para PESQUISAR O STATUS DO PAGAMENTO daquele cliente (cpf)
driver = webdriver.Chrome()
driver.get("https://consultcpf-devaprender.netlify.app")


for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha
    # print(nome, valor, cpf, vencimento)  # Descomente se quiser visualizar os dados



# O CPF será inserido atravez de seu imput XPATH //tag[@atributo='valor']
campo_pesquisa = driver.find_element(By.XPATH,'//input[@id="cpfInput"]')
sleep(1)
# Escrevendo o cpf
campo_pesquisa.send_keys(cpf)
sleep(1)


# 3 - Verificar se está em dia ou atrasado
botao_pesquisa = driver.find_element(By.XPATH,'//button[@class="btn btn-custom btn-lg btn-block mt-3"]')
sleep(1)
botao_pesquisa.click()
sleep(4)


status = driver.find_element(By.XPATH,"//span[@id='statusLabel']")
# Obter o texto do elemento
if status.text == "em dia":
    # 4 - Se estiver "EM DIA", pegar a DATA e METODO de pagamento 
    data_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentDate']")
    metodo_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentMethod']")
    pagina_fechamento.append([nome, valor, cpf, vencimento, "em dia","xxx","xxx"])
else:
    # 5 - Caso contrário (se estiver atrasado), colocar o status como pendente
    planilha_fechamento = openpyxl.load_workbook('D:\\MEUSSITESEPROJETOS\\Projetos\\Projetos-Python\\Projects4\\planilha_fechamento.xlsx') 
    pagina_fechamento = planilha_fechamento['Sheet1']


# Insere dados na planilha
pagina_fechamento.append([nome, valor, cpf, vencimento, "pendente"])





# 4 - Se estiver "EM DIA", pegar a DATA e METODO de pagamento 
# 5 - Caso contrário (se estiver atrasad), colocar o status como pendente
# 6 - Inserir essas novas informações (nome, valor, cpf e vencimento - status) e caso esteja em dia, 
# DATA DE PAGAMENTO, E METODO PAGAMENTO (cartão ou boleto) e, uma nova planilha



# Aguarde e inspecione o site manualmente
input("Pressione Enter para fechar o navegador...")

# Fechar o navegador
driver.quit()



# Não fechar o navegador automaticamente após 5 segundos
# Use a função `WebDriverWait` para aguardar até que o elemento esteja presente
# ou uma interação específica esteja concluída

# Aguarde a página processar a pesquisa
sleep(5)  # Ou use WebDriverWait para esperar um elemento específico

# Lógica adicional para extrair e manipular dados da página vai aqui






